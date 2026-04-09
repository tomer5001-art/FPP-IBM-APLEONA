#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FPP Web App - Facility Project Proposal
Web interface for generating IBM FPP Excel reports
"""

import os, io, smtplib, json, glob, base64
import urllib.request, urllib.error
from datetime import date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from PIL import Image as PILImage
import streamlit as st
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="FPP Generator – IBM Apleona",
    page_icon="🏢",
    layout="wide",
)

# ── Constants ─────────────────────────────────────────────────────────────────
CONTRACT_NUMBERS = (
    "MSA-CW3790072\n"
    "FM SOW-CW3790289\n"
    "PA FM SOW-CW3796451-IBM Israel Ltd\n"
    "PA FM SOW-CW3796611-IBM Israel S&T Ltd"
)
PREPARED_BY      = "Tomer Cohen"
FEE              = 0.06
SAVE_DIR         = r"C:\Users\TomerCohen\Apleona Group\Apleona Israel - General\All Israel Clients\IBM\IBM 2026\FPP TO IBM-2026\APPLICETION"
SAVE_DIR2        = r"C:\Users\TomerCohen\Apleona Group\Apleona Israel - General\Claude code"
GITHUB_OWNER     = "Tomer5001"
GITHUB_REPO      = "fpp-ibm-apleona"
GITHUB_HIST_PATH = "history"

LABOR_ROLES = [
    {"title": "Small Job Coordinator Sr",         "st_rate": 358.49, "ot_rate": 537.74},
    {"title": "Small Job Coordinator",            "st_rate": 286.79, "ot_rate": 430.19},
    {"title": "Facilities Engineer Sr",           "st_rate": 269.81, "ot_rate": 404.72},
    {"title": "Facilities Engineer",              "st_rate": 248.11, "ot_rate": 372.17},
    {"title": "Facilities Technician",            "st_rate": 162.26, "ot_rate": 243.39},
    {"title": "Facilities Administrator",         "st_rate": 130.19, "ot_rate": 195.29},
    {"title": "Space Administrator",              "st_rate": 216.04, "ot_rate": 324.06},
    {"title": "Handyman/Porter",                  "st_rate": 118.87, "ot_rate": 178.31},
    {"title": "Workplace Experience Manager",     "st_rate": 301.89, "ot_rate": 452.84},
    {"title": "Workplace Experience Coordinator", "st_rate": 301.89, "ot_rate": 452.84},
    {"title": "Concierge",                        "st_rate":  85.85, "ot_rate": 128.78},
]

SITES = ["PTK", "Haifa", "Shahar", "Be'er Sheva"]

# ── Styling ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    [data-testid="stAppViewContainer"] { direction: rtl; }
    .main-title {
        background: linear-gradient(135deg, #1F3864, #2E75B6);
        color: white; padding: 28px 32px; border-radius: 12px;
        margin-bottom: 28px; text-align: center;
    }
    .main-title h1 { margin: 0; font-size: 2rem; }
    .main-title p  { margin: 6px 0 0; opacity: .85; font-size: 1rem; }
    .section-card {
        background: #f8fafc; border: 1px solid #dce6f1;
        border-radius: 10px; padding: 22px 26px; margin-bottom: 22px;
    }
    .section-title {
        color: #1F3864; font-size: 1.1rem; font-weight: 700;
        border-bottom: 2px solid #2E75B6; padding-bottom: 8px; margin-bottom: 16px;
    }
    .history-card {
        background: #f0f7ff; border: 1px solid #b3d1f0;
        border-radius: 10px; padding: 18px 22px; margin-bottom: 22px;
    }
    .stButton > button {
        background: linear-gradient(135deg, #1F3864, #2E75B6);
        color: white; border: none; padding: 14px 40px;
        font-size: 1.1rem; border-radius: 8px; width: 100%;
        font-weight: 600; cursor: pointer;
    }
    .stButton > button:hover { opacity: .9; }
    label { font-weight: 600 !important; color: #1F3864 !important; }
</style>
""", unsafe_allow_html=True)

# ── GitHub history helpers ────────────────────────────────────────────────────
def _gh_headers(token):
    return {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json",
        "Content-Type": "application/json",
    }

def github_save_json(payload, filename, token):
    url     = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{GITHUB_HIST_PATH}/{filename}"
    content = base64.b64encode(json.dumps(payload, ensure_ascii=False, indent=2).encode()).decode()
    sha     = None
    try:
        req = urllib.request.Request(url, headers=_gh_headers(token))
        with urllib.request.urlopen(req) as r:
            sha = json.loads(r.read()).get("sha")
    except Exception:
        pass
    body = {"message": f"FPP history: {filename}", "content": content}
    if sha:
        body["sha"] = sha
    try:
        req = urllib.request.Request(url, data=json.dumps(body).encode(),
                                     headers=_gh_headers(token), method="PUT")
        urllib.request.urlopen(req)
        return True
    except Exception:
        return False

def github_load_history(token):
    url     = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{GITHUB_HIST_PATH}"
    records = []
    try:
        req = urllib.request.Request(url, headers=_gh_headers(token))
        with urllib.request.urlopen(req) as r:
            files = json.loads(r.read())
        for f in sorted(files, key=lambda x: x["name"], reverse=True):
            if f["name"].endswith(".json"):
                try:
                    req2 = urllib.request.Request(f["download_url"])
                    with urllib.request.urlopen(req2) as r2:
                        records.append(json.loads(r2.read()))
                except Exception:
                    pass
    except Exception:
        pass
    return records

# ── Local history helpers ─────────────────────────────────────────────────────
def local_save_json(payload, filename):
    try:
        os.makedirs(SAVE_DIR, exist_ok=True)
        with open(os.path.join(SAVE_DIR, filename), "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def local_load_history():
    records = []
    try:
        for path in sorted(glob.glob(os.path.join(SAVE_DIR, "*.json")), reverse=True):
            try:
                with open(path, encoding="utf-8") as f:
                    records.append(json.load(f))
            except Exception:
                pass
    except Exception:
        pass
    return records

# ── Email sender ─────────────────────────────────────────────────────────────
def send_email(project_name, site, filename, excel_bytes, secrets):
    try:
        sender     = secrets.get("EMAIL_SENDER", "")
        password   = secrets.get("EMAIL_PASSWORD", "")
        recipients = ["tomer.cohen2@ibm.com", "jonatan.ben.sudai@ibm.com"]
        if not sender or not password:
            return False, "פרטי מייל חסרים ב-Secrets"
        msg = MIMEMultipart()
        msg["From"]    = sender
        msg["To"]      = ", ".join(recipients)
        msg["Subject"] = f"נוצר FPP חדש – {project_name}"
        body = f"""שלום תומר,\n\nנוצר FPP חדש במערכת Apleona FPP Generator.\n\nפרטים:\n• שם פרויקט: {project_name}\n• אתר: {site}\n• תאריך: {date.today().strftime('%d/%m/%Y')}\n• שם קובץ: {filename}\n\nהקובץ מצורף למייל זה.\n\nבברכה,\nמערכת FPP Generator – Apleona Israel\n"""
        msg.attach(MIMEText(body, "plain", "utf-8"))
        part = MIMEBase("application", "octet-stream")
        part.set_payload(excel_bytes)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
        msg.attach(part)
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender, password)
            server.sendmail(sender, recipients, msg.as_string())
        return True, ""
    except Exception as e:
        return False, str(e)

# ── Translation ───────────────────────────────────────────────────────────────
def translate(client, text):
    if not text or not any("\u0590" <= c <= "\u05FF" for c in text):
        return text
    msg = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=1024,
        messages=[{"role": "user", "content": (
            "Translate the following Hebrew text to professional English "
            "suitable for a formal facility management proposal submitted to IBM. "
            "Preserve line breaks. Return ONLY the translation.\n\n" + text
        )}]
    )
    return msg.content[0].text.strip()

# ── Excel helpers ─────────────────────────────────────────────────────────────
def _b():
    t = Side(style="thin", color="B8C4CE")
    return Border(left=t, right=t, top=t, bottom=t)

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=9): return Font(name="Arial", bold=bold, color=color, size=size)
def _align(h="left", v="center", wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def sc(cell, value=None, font=None, fill=None, align=None, fmt=None):
    if value is not None: cell.value = value
    if font:  cell.font      = font
    if fill:  cell.fill      = fill
    if align: cell.alignment = align
    if fmt:   cell.number_format = fmt
    cell.border = _b()

def generate_excel(data) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col, w in zip("ABCDEFG", [46, 15, 12, 12, 17, 8, 17]):
        ws.column_dimensions[col].width = w
    br = _b()
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:G1")
    sc(ws["A1"], "Facility Project Proposal",
       font=_font(bold=True, color="FFFFFF", size=14),
       fill=_fill("1F3864"), align=_align("center"))
    lf, lfl = _font(bold=True), _fill("F8F8F8")
    vf, vfl = _font(), _fill("F8F8F8")
    def hrow(row, label, value, height=16, va=None):
        ws.row_dimensions[row].height = height
        ws.merge_cells(f"B{row}:G{row}")
        sc(ws[f"A{row}"], label, font=lf, fill=lfl, align=_align())
        sc(ws[f"B{row}"], value, font=vf, fill=vfl, align=va or _align())
    hrow(2, "Project Name:", data["project_name"])
    hrow(3, "Site:",         data["site"])
    hrow(4, "Contract #:",   CONTRACT_NUMBERS, height=70, va=_align("left","top"))
    hrow(5, "Revision Date:",data["revision_date"])
    ws["B5"].number_format = "DD/MM/YYYY"
    hrow(6, "Fixed Price or Estimate/Not to Exceed (NTE)", "Fixed")
    hrow(7, "Proposal Prepared by:", PREPARED_BY)
    ws.row_dimensions[8].height = 6
    ws.row_dimensions[9].height = 18
    ws.merge_cells("A9:G9")
    sc(ws["A9"], "Scope of Work:", font=_font(bold=True), fill=_fill("CFE2F3"), align=_align())
    ws.merge_cells("A10:G11")
    sc(ws["A10"], data["scope_of_work"], font=vf, fill=vfl, align=_align("left","top"))
    ws.row_dimensions[10].height = 65
    ws.row_dimensions[11].height = 65
    ws.row_dimensions[12].height = 18
    ws.merge_cells("A12:G12")
    sc(ws["A12"], "Financial Proposal:", font=_font(bold=True), fill=_fill("CFE2F3"), align=_align())
    ws.row_dimensions[13].height = 28
    thf, thfl = _font(bold=True, color="FFFFFF"), _fill("2E75B6")
    for col, hdr in enumerate(["Description","Unit Price","Quantity","UoM","Michlol Net ILS","Fee","Apleona net ILS"],1):
        sc(ws.cell(row=13,column=col,value=hdr), font=thf, fill=thfl, align=_align("center"))
    alt, base = _fill("EBF3FB"), _fill("F8F8F8")
    cur = 14
    mat_items = data["items"]
    for i, item in enumerate(mat_items):
        ws.row_dimensions[cur].height = 18
        f = alt if i % 2 == 0 else base
        michlol = round(item["unit_price"] * item["quantity"] * 1.05, 2)
        apleona = round(michlol * (1 + FEE), 2)
        sc(ws.cell(row=cur,column=1), item["description"],   font=vf,              fill=f, align=_align())
        sc(ws.cell(row=cur,column=2), item["unit_price"],    font=_font(bold=True), fill=f, align=_align("right",wrap=False), fmt="#,##0.00")
        sc(ws.cell(row=cur,column=3), item["quantity"],      font=vf,              fill=f, align=_align("center"))
        sc(ws.cell(row=cur,column=4), item["uom"],           font=vf,              fill=f, align=_align("center"))
        sc(ws.cell(row=cur,column=5), michlol,               font=vf,              fill=f, align=_align("right",wrap=False), fmt="#,##0.00")
        sc(ws.cell(row=cur,column=6), FEE,                   font=vf,              fill=f, align=_align("center"), fmt="0%")
        sc(ws.cell(row=cur,column=7), apleona,               font=vf,              fill=f, align=_align("right",wrap=False), fmt="#,##0.00")
        cur += 1
    active_labor = [r for r in data["labor_roles"] if r["st_hours"] > 0 or r["ot_hours"] > 0]
    if active_labor:
        ws.row_dimensions[cur].height = 16
        ws.merge_cells(f"A{cur}:G{cur}")
        sc(ws.cell(row=cur,column=1), "Self Performed Labor",
           font=_font(bold=True), fill=_fill("D6E4F0"), align=_align())
        cur += 1
        for i, role in enumerate(active_labor):
            ws.row_dimensions[cur].height = 18
            f = alt if i % 2 == 0 else base
            labor_cost    = round(role["st_hours"] * role["st_rate"] + role["ot_hours"] * role["ot_rate"], 2)
            labor_apleona = round(labor_cost * (1 + FEE), 2)
            total_hours   = role["st_hours"] + role["ot_hours"]
            sc(ws.cell(row=cur,column=1), role["title"],   font=_font(bold=True), fill=f, align=_align())
            sc(ws.cell(row=cur,column=2), labor_cost,      font=_font(bold=True), fill=f, align=_align("right",wrap=False), fmt="#,##0.00")
            sc(ws.cell(row=cur,column=3), total_hours,     font=vf,               fill=f, align=_align("center"))
            sc(ws.cell(row=cur,column=4), "Hours",         font=vf,               fill=f, align=_align("center"))
            sc(ws.cell(row=cur,column=5), labor_cost,      font=vf,               fill=f, align=_align("right",wrap=False), fmt="#,##0.00")
            sc(ws.cell(row=cur,column=6), FEE,             font=vf,               fill=f, align=_align("center"), fmt="0%")
            sc(ws.cell(row=cur,column=7), labor_apleona,   font=vf,               fill=f, align=_align("right",wrap=False), fmt="#,##0.00")
            cur += 1
    total_michlol = sum(round(it["unit_price"]*it["quantity"]*1.05,2) for it in mat_items if it["unit_price"]) + \
                    sum(round(r["st_hours"]*r["st_rate"]+r["ot_hours"]*r["ot_rate"],2) for r in active_labor)
    total_apleona = sum(round(it["unit_price"]*it["quantity"]*1.05*(1+FEE),2) for it in mat_items if it["unit_price"]) + \
                    sum(round((r["st_hours"]*r["st_rate"]+r["ot_hours"]*r["ot_rate"])*(1+FEE),2) for r in active_labor)
    tf, tfl = _font(bold=True), _fill("BDD7EE")
    ws.row_dimensions[cur].height = 22
    sc(ws.cell(row=cur,column=1), "Total net", font=tf, fill=tfl, align=_align())
    for col in range(2,8):
        ws.cell(row=cur,column=col).fill = tfl
        ws.cell(row=cur,column=col).border = br
    sc(ws.cell(row=cur,column=5), total_michlol, font=tf, fill=tfl, align=_align("right",wrap=False), fmt="#,##0.00")
    sc(ws.cell(row=cur,column=7), total_apleona, font=tf, fill=tfl, align=_align("right",wrap=False), fmt="#,##0.00")
    cur += 1
    ws.row_dimensions[cur].height = 18
    ws.merge_cells(f"A{cur}:G{cur}")
    sc(ws.cell(row=cur,column=1), "Clarifications / Assumptions:",
       font=_font(bold=True), fill=_fill("CFE2F3"), align=_align())
    cur += 1
    ws.row_dimensions[cur].height = 6
    cur += 1
    ws.merge_cells(f"A{cur}:G{cur+3}")
    sc(ws.cell(row=cur,column=1), data["clarifications"] or " ",
       font=vf, fill=vfl, align=_align("left","top"))
    for r in range(cur, cur+4): ws.row_dimensions[r].height = 20
    cur += 4
    ws.row_dimensions[cur].height = 18
    ws.merge_cells(f"A{cur}:G{cur}")
    sc(ws.cell(row=cur,column=1), "FM Provider Self Performed Labor Breakdown:",
       font=_font(bold=True), fill=_fill("CFE2F3"), align=_align())
    cur += 1
    ws.row_dimensions[cur].height = 18
    for col, hdr in enumerate(["Job Title","ST Hours","ST Rate","ST Total","OT Hours","OT Rate","OT Total"],1):
        sc(ws.cell(row=cur,column=col,value=hdr), font=thf, fill=thfl, align=_align("center"))
    cur += 1
    st_total_all = ot_total_all = 0
    for i, role in enumerate(data["labor_roles"]):
        ws.row_dimensions[cur].height = 14
        f = alt if i % 2 == 0 else base
        st_total = round(role["st_hours"] * role["st_rate"], 2)
        ot_total = round(role["ot_hours"] * role["ot_rate"], 2)
        st_total_all += st_total; ot_total_all += ot_total
        sc(ws.cell(row=cur,column=1), role["title"],    font=_font(bold=True), fill=f, align=_align())
        sc(ws.cell(row=cur,column=2), role["st_hours"], font=_font(bold=True), fill=f, align=_align("center"))
        sc(ws.cell(row=cur,column=3), role["st_rate"],  font=vf,              fill=f, align=_align("center"), fmt="#,##0.00")
        sc(ws.cell(row=cur,column=4), st_total,         font=vf,              fill=f, align=_align("right",wrap=False), fmt="#,##0.00")
        sc(ws.cell(row=cur,column=5), role["ot_hours"], font=_font(bold=True), fill=f, align=_align("center"))
        sc(ws.cell(row=cur,column=6), role["ot_rate"],  font=vf,              fill=f, align=_align("center"), fmt="#,##0.00")
        sc(ws.cell(row=cur,column=7), ot_total,         font=vf,              fill=f, align=_align("right",wrap=False), fmt="#,##0.00")
        cur += 1
    ws.row_dimensions[cur].height = 18
    sc(ws.cell(row=cur,column=1), "TOTAL", font=tf, fill=tfl, align=_align())
    for col in range(2,8):
        ws.cell(row=cur,column=col).fill = tfl
        ws.cell(row=cur,column=col).border = br
    sc(ws.cell(row=cur,column=4), round(st_total_all,2), font=tf, fill=tfl, align=_align("right",wrap=False), fmt="#,##0.00")
    sc(ws.cell(row=cur,column=7), round(ot_total_all,2), font=tf, fill=tfl, align=_align("right",wrap=False), fmt="#,##0.00")
    cur += 1
    ws.print_area = f"A1:G{cur-1}"
    ws.page_setup.orientation="landscape"; ws.page_setup.fitToPage=True
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    if data.get("images"):
        ps = wb.create_sheet("Photos & Attachments")
        ps.column_dimensions["A"].width = 12
        ps.column_dimensions["B"].width = 80
        sc(ps["A1"], "Attachments", font=_font(bold=True, color="FFFFFF", size=11),
           fill=_fill("1F3864"), align=_align("center"))
        ps.merge_cells("A1:B1")
        ps.row_dimensions[1].height = 24
        cur_row = 2
        for idx, (fname, img_bytes) in enumerate(data["images"], 1):
            ps.row_dimensions[cur_row].height = 16
            sc(ps.cell(row=cur_row,column=1), f"File {idx}:", font=_font(bold=True), fill=_fill("CFE2F3"), align=_align())
            sc(ps.cell(row=cur_row,column=2), fname,          font=_font(),           fill=_fill("F8F8F8"), align=_align())
            cur_row += 1
            try:
                pil = PILImage.open(io.BytesIO(img_bytes))
                max_w = 600
                if pil.width > max_w:
                    ratio = max_w / pil.width
                    pil = pil.resize((max_w, int(pil.height * ratio)), PILImage.LANCZOS)
                img_buf = io.BytesIO()
                fmt = pil.format or "PNG"
                if fmt.upper() == "JPG": fmt = "JPEG"
                pil.save(img_buf, format=fmt); img_buf.seek(0)
                xl_img = XLImage(img_buf); xl_img.anchor = f"B{cur_row}"
                ps.add_image(xl_img)
                img_rows = max(1, int(pil.height / 14))
                for r in range(cur_row, cur_row + img_rows): ps.row_dimensions[r].height = 14
                cur_row += img_rows + 1
            except Exception:
                sc(ps.cell(row=cur_row,column=2), "(non-image file — see filename above)",
                   font=_font(), fill=_fill("F8F8F8"), align=_align())
                cur_row += 2
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="main-title">
    <h1>🏢 Facility Project Proposal</h1>
    <p>Apleona Israel · IBM Account · FPP Generator</p>
</div>
""", unsafe_allow_html=True)

if "num_cost_rows" not in st.session_state: st.session_state.num_cost_rows = 2
if "fpp_load"      not in st.session_state: st.session_state.fpp_load = {}
L = st.session_state.fpp_load

gh_token = st.secrets.get("GITHUB_TOKEN", os.environ.get("GITHUB_TOKEN", ""))

# ── Section 0: History ────────────────────────────────────────────────────────
history = github_load_history(gh_token) if gh_token else local_load_history()

if history:
    st.markdown('<div class="history-card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">📂 טען FPP קודם לעריכה</div>', unsafe_allow_html=True)
    options = ["— בחר FPP קודם —"] + [
        f"{r.get('created_date','?')}  |  {r.get('project_name_he','?')}  |  {r.get('site','?')}"
        for r in history
    ]
    chosen = st.selectbox("FPP קודמים", options, label_visibility="collapsed")
    if chosen != options[0]:
        idx = options.index(chosen) - 1
        rec = history[idx]
        if st.button("📂 טען לטופס"):
            st.session_state.fpp_load = rec
            st.session_state.num_cost_rows = max(2, len(rec.get("cost_items", [])))
            st.rerun()
    if L:
        if st.button("🗑️ נקה טופס"):
            st.session_state.fpp_load = {}
            st.session_state.num_cost_rows = 2
            st.rerun()
        st.info(f"✏️ עורך: **{L.get('project_name_he','')}** ({L.get('site','')}  {L.get('created_date','')})")
    st.markdown('</div>', unsafe_allow_html=True)

# ── Section 1: Project Details ────────────────────────────────────────────────
st.markdown('<div class="section-card"><div class="section-title">📋 פרטי הפרויקט</div>', unsafe_allow_html=True)
col1, col2 = st.columns(2)
with col1:
    project_name = st.text_input("שם הפרויקט *", value=L.get("project_name_he",""),
        placeholder="לדוגמה: החלפת מזגן במשרד PTK")
with col2:
    site_default = L.get("site", SITES[0])
    site = st.selectbox("אתר *", SITES, index=SITES.index(site_default) if site_default in SITES else 0)
st.markdown('</div>', unsafe_allow_html=True)

# ── Section 2: Scope of Work ──────────────────────────────────────────────────
st.markdown('<div class="section-card"><div class="section-title">📝 מהות העבודה</div>', unsafe_allow_html=True)
st.caption("תאר את העבודה בחופשיות בעברית — יתורגם לאנגלית מקצועית אוטומטית")
scope_he = st.text_area("תיאור העבודה *", value=L.get("scope_he",""), height=140,
    placeholder="לדוגמה:\nהחלפת יחידת מיזוג אוויר בקומה 3\nפירוק היחידה הישנה ופינוי לאתר פסולת מאושר\nהתקנת יחידה חדשה כולל בדיקות הרצה")
st.markdown('</div>', unsafe_allow_html=True)

# ── Section 3: Costs ──────────────────────────────────────────────────────────
st.markdown('<div class="section-card"><div class="section-title">💰 עלויות</div>', unsafe_allow_html=True)
st.caption("המערכת תוסיף 6% אוטומטית ותחשב סה\"כ")
loaded_items = L.get("cost_items", [])
cost_items = []
for i in range(st.session_state.num_cost_rows):
    prev = loaded_items[i] if i < len(loaded_items) else {}
    cols = st.columns([3, 1, 1.5])
    with cols[0]:
        desc = st.text_input(f"תיאור פריט {i+1}", key=f"desc_{i}",
            value=prev.get("description_he",""),
            placeholder="לדוגמה: אספקת מזגן 24BTU" if i == 0 else "")
    with cols[1]:
        qty = st.number_input("כמות", key=f"qty_{i}", min_value=1, step=1,
            value=int(prev.get("quantity", 1)))
    with cols[2]:
        price = st.number_input("מחיר יחידה (₪)", key=f"price_{i}",
            min_value=0.0, step=100.0, format="%.2f",
            value=float(prev.get("unit_price", 0.0)))
    if desc:
        cost_items.append({"description_he": desc, "unit_price": price, "quantity": qty})
if st.button("➕ הוסף שורה"):
    st.session_state.num_cost_rows += 1
    st.rerun()
if cost_items:
    subtotal = sum(i["unit_price"] * i["quantity"] for i in cost_items)
    total_with_fee = subtotal * 1.05 * 1.06
    st.markdown(f"""
    <div style="background:#ddeeff;border-radius:8px;padding:12px 18px;margin-top:12px;text-align:left">
        <b>תת-סה"כ:</b> ₪{subtotal:,.2f} &nbsp;|&nbsp;
        <b>סה"כ כולל 5% מכלול + 6% עמלה:</b>
        <span style="color:#1F3864;font-size:1.1em">₪{total_with_fee:,.2f}</span>
    </div>""", unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

# ── Section 4: Labor ──────────────────────────────────────────────────────────
st.markdown('<div class="section-card"><div class="section-title">👷 שעות עובד נוסף</div>', unsafe_allow_html=True)
st.caption("הכנס שעות לתפקידים הרלוונטיים בלבד (0 = לא רלוונטי)")
loaded_labor = L.get("labor", {})
labor_data = []
for role in LABOR_ROLES:
    prev_l   = loaded_labor.get(role["title"], {})
    prev_st  = int(prev_l.get("st_hours", 0))
    prev_ot  = int(prev_l.get("ot_hours", 0))
    prev_hrs = prev_st + prev_ot
    col1, col2, col3 = st.columns([2.5, 1, 1.8])
    with col1:
        st.markdown(f"<span style='font-weight:600'>{role['title']}</span>", unsafe_allow_html=True)
    with col2:
        hours = st.number_input("שעות", key=f"h_{role['title']}",
            min_value=0, step=1, value=prev_hrs, label_visibility="collapsed")
    with col3:
        if hours > 0:
            default_shift = "מחוץ לשעות (OT)" if prev_ot > 0 and prev_st == 0 else "בשעות העבודה (ST)"
            shift = st.radio("משמרת", ["בשעות העבודה (ST)", "מחוץ לשעות (OT)"],
                key=f"s_{role['title']}", horizontal=True,
                index=0 if default_shift == "בשעות העבודה (ST)" else 1,
                label_visibility="collapsed")
            st_h = int(hours) if "ST" in shift else 0
            ot_h = int(hours) if "OT" in shift else 0
        else:
            st.markdown(""); st_h = ot_h = 0
    labor_data.append({**role, "st_hours": st_h, "ot_hours": ot_h})
st.markdown('</div>', unsafe_allow_html=True)

# ── Section 5: Clarifications ─────────────────────────────────────────────────
st.markdown('<div class="section-card"><div class="section-title">📌 הבהרות והנחות</div>', unsafe_allow_html=True)
clarifications_he = st.text_area("הבהרות/הנחות (אופציונלי)", value=L.get("clarifications_he",""),
    height=100, placeholder="לדוגמה: המחיר אינו כולל עבודות בנייה אזרחית")
st.markdown('</div>', unsafe_allow_html=True)

# ── Section 6: File Attachments ───────────────────────────────────────────────
st.markdown('<div class="section-card"><div class="section-title">📎 קבצים מצורפים</div>', unsafe_allow_html=True)
st.caption("ניתן לצרף תמונות או קבצים — יצורפו לאקסל בגיליון נפרד")
uploaded_files = st.file_uploader("העלה תמונות / קבצים", accept_multiple_files=True,
    type=["png","jpg","jpeg","gif","bmp","webp","pdf"], label_visibility="collapsed")
if uploaded_files:
    st.success(f"✓ {len(uploaded_files)} קובץ/ים הועלו")
    for f in uploaded_files: st.caption(f"📄 {f.name}")
st.markdown('</div>', unsafe_allow_html=True)

# ── Generate ──────────────────────────────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
if st.button("⚡ צור קובץ FPP"):
    if not project_name:
        st.error("אנא הכנס שם פרויקט")
    elif not scope_he:
        st.error("אנא הכנס תיאור עבודה")
    elif not cost_items:
        st.error("אנא הכנס לפחות פריט עלות אחד")
    else:
        with st.spinner("מתרגם לאנגלית ומייצר קובץ Excel..."):
            try:
                api_key = st.secrets.get("ANTHROPIC_API_KEY", os.environ.get("ANTHROPIC_API_KEY"))
                client  = anthropic.Anthropic(api_key=api_key)
                project_name_en = translate(client, project_name)
                scope_en        = translate(client, scope_he)
                clari_en        = translate(client, clarifications_he) if clarifications_he else ""
                items_en = []
                for item in cost_items:
                    desc_en = translate(client, item["description_he"])
                    items_en.append({
                        "description": desc_en,
                        "unit_price":  item["unit_price"],
                        "quantity":    item["quantity"],
                        "uom":         "Lump Sum" if item["quantity"] == 1 else "Units",
                        "fee":         FEE,
                    })
                while len(items_en) < 10:
                    items_en.append({"description":"","unit_price":0,"quantity":0,"uom":"-","fee":FEE})
                images = [(f.name, f.read()) for f in uploaded_files] if uploaded_files else []
                excel_bytes = generate_excel({
                    "project_name":   project_name_en,
                    "site":           site,
                    "revision_date":  date.today(),
                    "scope_of_work":  scope_en,
                    "items":          items_en,
                    "clarifications": clari_en,
                    "labor_roles":    labor_data,
                    "images":         images,
                })
                safe     = "".join(c if c.isalnum() or c in " _-" else "_" for c in project_name_en)[:40].strip()
                filename = f"FPP_{safe}_{site}_{date.today().strftime('%Y%m%d')}.xlsx"
                for sd in [SAVE_DIR, SAVE_DIR2]:
                    try:
                        os.makedirs(sd, exist_ok=True)
                        with open(os.path.join(sd, filename), "wb") as f:
                            f.write(excel_bytes)
                    except Exception:
                        pass
                json_payload = {
                    "project_name_he":   project_name,
                    "site":              site,
                    "scope_he":          scope_he,
                    "clarifications_he": clarifications_he,
                    "cost_items":        cost_items,
                    "labor":             {r["title"]: {"st_hours": r["st_hours"], "ot_hours": r["ot_hours"]} for r in labor_data},
                    "created_date":      date.today().strftime("%Y-%m-%d"),
                    "filename":          filename,
                }
                json_filename = filename.replace(".xlsx", ".json")
                if gh_token:
                    github_save_json(json_payload, json_filename, gh_token)
                else:
                    local_save_json(json_payload, json_filename)
                send_email(project_name_en, site, filename, excel_bytes, st.secrets)
                st.download_button(
                    label="📥 הורד קובץ Excel",
                    data=excel_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_btn",
                )
                st.markdown("""
                <script>
                setTimeout(function() {
                    try {
                        var ctx = new (window.AudioContext || window.webkitAudioContext)();
                        var o = ctx.createOscillator(); var g = ctx.createGain();
                        o.connect(g); g.connect(ctx.destination);
                        o.type = 'sine'; o.frequency.value = 880;
                        g.gain.setValueAtTime(0.15, ctx.currentTime);
                        g.gain.exponentialRampToValueAtTime(0.0001, ctx.currentTime + 0.3);
                        o.start(ctx.currentTime); o.stop(ctx.currentTime + 0.3);
                    } catch(e) {}
                    var btns = window.parent.document.querySelectorAll('[data-testid="stDownloadButton"] button');
                    if (btns.length > 0) { btns[0].click(); }
                }, 1000);
                </script>
                """, unsafe_allow_html=True)
            except Exception as e:
                st.error(f"שגיאה: {e}")
